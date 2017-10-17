# Module Imports
import csv
import re
import json
import os
import collections
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fuzzywuzzy import process


def read_schema(filename):
    # Function to read the schema files (which define where to look for all the data) into lists

    csvcontents = []

    try:
        # Attempt to open & parse the schema
        with open(filename, newline='') as csvfile:

            csvobj = csv.reader(csvfile, delimiter=',', quotechar='|')

            for row in csvobj:
                csvcontents.append(row)

        return csvcontents

    except FileNotFoundError:
        message = 'Could Not Find Schema'
        print( message + ' ' + filename)
        write_to_log(['Error', message, filename, 'N/A'])
        exit() # Cant do much without the schema!

def get_schema(w_sheet):
    # Function which takes a worksheet and returns the appropriate schema

    # Look in Cell B5 which contains different things for each sheet type
    sheet_indicator = w_sheet['B5'].value

    if sheet_indicator == 'EPI: CHILDREN':
        schema = read_schema('../Schema/Immunisation Schema_01.csv')
        schema_id = 'Immunisation Schema_01'
        s_type='Immunisation'

    elif sheet_indicator == 'OUTPATIENT - CURATIVE SERVICES':
        schema = read_schema('../Schema/OPD Schema_01.csv')
        schema_id = 'OPD Schema_01'
        s_type = 'OPD'

    elif sheet_indicator == 'SAFE MOTHERHOOD PROGRAMME':
        schema = read_schema('../Schema/Motherhood Schema_01.csv')
        schema_id = 'Motherhood Schema_01'
        s_type = 'Motherhood'

    else:
        print('Error - Sheet format not identified')

    return s_type, schema, schema_id


def year_from_fname(fname):
# Use regular expressions to extract *20XX* year from a filename where XX is two numbers
    try:
        re_year = re.search(r".*(20[0-9]{2}).*", fname).group(1)
    except AttributeError:
        message = 'Year Not Found in Filename'
        print('\n' + message + ': ' + fname)
        write_to_log(['Error', message, fname, 'N/A' ])
        re_year = 0

    year = int(re_year)

    return year


def schema2header(schema):
    # function to format ouput of read_schema to some nice headers

    headers = [] #Empty list to fill with headers

    for item in schema:
        # Schema consists of a list of lists, join the items in the inner lists using newline chars to generate headers
        headers.append('\n'.join(item[1:]))

    return headers


def init_master_csv():
    # Initialise CSV file, writing headers from schema
    with open('../Output/master_out.csv', 'w', newline='') as csvfile:
        wr = csv.writer(csvfile, dialect='excel')
        wr.writerow(['Compiled healthcare data from files in \'Input\'',])

        # Column Headings for Non-Schema Rows
        col_headings = ['Year', 'Month', 'Region', 'Clinic', ' ',                     # Items ID'ing Data
                        'Clinic Type', 'District', 'Donor',  'Est.Pop.', ' ']         # Items from Lookup

        # Read in schema as lists
        imn_schema = schema2header(read_schema('../Schema/Immunisation Schema_01.csv'))
        opd_schema = schema2header(read_schema('../Schema/OPD Schema_01.csv'))
        mhd_schema = schema2header(read_schema('../Schema/Motherhood Schema_01.csv'))

        # Combine the lists& column headings
        combined_schema = col_headings + imn_schema + opd_schema + mhd_schema

        wr.writerow(combined_schema)

    return col_headings, imn_schema, opd_schema, mhd_schema


def append_data_rows(rows):
    # Function to append rows to the master_out csv initialised by init_mastercsv

    with open('../Output/master_out.csv', 'a', newline='') as csvfile:
        wr = csv.writer(csvfile, dialect='excel')
        for row in rows:
            wr.writerow(row)


def prune(rows, n):

    pruned_rows = []

    # Iterate over input rows
    for row in rows:

        # iterate over the items in each row
        for item in row[n:-1]:

            # if an item is found in the row that is not none and not '' then the row is appended
            if (item is not None) and (item is not ''):
                pruned_rows.append(row)
                break  # break out of item for loop, back into row for loop

    return pruned_rows


def gen_ordered_dict(lst):

    # Insert each set of items into an ordered dictionary
    od = collections.OrderedDict()
    for item in lst:
        od[item[0]] = item

    return od


def clean_rows(rows, places_od):
    #
    # run over rows,
    #
    # if the row is not in the 'correct' list of names,
    #     iterate over each item in our dictionaries of alternatives
    #
    #     if it is in one of them
    #         change its name to the 'correct' one (i.e. the key)
    #
    #   will now be left with duplicate rows - same year, month, place
    #
    cleaned_rows=[]

    return cleaned_rows

# Need a way of grouping alternative strings for months
months = [
            ['Jan', 'jan', 'January', 'january'],
            ['Feb', 'feb', 'February', 'february', 'Febuary', 'febuary'],
            ['Mar', 'mar', 'March', 'march'],
            ['Apr', 'apr', 'April', 'april'],
            ['May', 'may'],
            ['Jun', 'jun', 'June', 'june'],
            ['Jul', 'jul'],
            ['Aug', 'aug', 'August', 'august'],
            ['Sep', 'sep', 'Sept', 'sept', 'september', 'September'],
            ['Oct', 'oct', 'October', 'october'],
            ['Nov', 'nov', 'November', 'november', 'Novem', 'novem', 'Novemb' 'novemb'],
            ['Dec', 'dec', 'December', 'december']]


def init_import_log():
    # Function to capture errors
    with open('../Output/import_log.csv', 'w', newline='') as logfile:
        wr = csv.writer(logfile, dialect='excel')
        wr.writerow(['Logfile listing data import errors'])
        wr.writerow(['Type','Message','File','Worksheet'])


def write_to_log(row):
    # Function to write rows to log csv file
    with open('../Output/import_log.csv','a', newline='') as logfile:
        wr = csv.writer(logfile, dialect='excel')
        wr.writerow(row)


def get_clinic_info():
    # point to xlsx containing info on all clinics
    clinic_info = '../Schema/20170804_clinic_list_master.xlsx'

    # Create a blank dictionary to store the clinic info in
    clinic_dict = dict()

    # Load the workbook and worksheet
    try:
        w_book = load_workbook(clinic_info, data_only=True)
        ws = w_book["Table"]

        # Get the column with the clinic unique names in (note 3:0 indexes from 3 to last row with info
        col_names = ws["G3:G0"]

        # for each clinic row
        for item in col_names:
            # Name Cell
            x = item[0]

            # use the row property of the cell we are on to index into the worksheet and extract items, storing them in a
            # dictionary
            new_dict = {
                "region": ws['A' + str(x.row)].value,
                "p_code": ws['B' + str(x.row)].value,
                "branch_coord": ws['C' + str(x.row)].value,
                "pop": ws['D' + str(x.row)].value,
                "name": ws['E' + str(x.row)].value,
                "status": ws['F' + str(x.row)].value,
                "lat": ws['H' + str(x.row)].value,
                "lon": ws['I' + str(x.row)].value,
                "coord_src": ws['J' + str(x.row)].value,
                "focal_pt": ws['K' + str(x.row)].value,
                "focal_no": ws['L' + str(x.row)].value,
                "funding": ws['N' + str(x.row)].value,
                "funding_end": ws['O' + str(x.row)].value,
                "district": ws['P' + str(x.row)].value,
                "info": ws['Q' + str(x.row)].value
            }

            # add this dictionary to the clinic dictionary with a key of the clinic name
            clinic_dict[x.value] = new_dict

    except FileNotFoundError:
        print('\nCould Not Find ' + clinic_info)
        write_to_log(['Error','Could Not Find Clinic Information', clinic_info, 'N/A'])

    return clinic_dict


def extract_clinic_info(all_clinics_dict, region, clinic):

    # Function using fuzzy matching to figure out which clinic in the in the all_clinics_dict produced by
    # get_clinic_info() is the best match for the provided (potentially misspelled) region & name strings

    # get the proper names for the clinics from the dictionary keys
    proper_names = all_clinics_dict.keys()

    # concatenate provided region & name with a space
    trial_name = region + " " + clinic

    try:
        # Select the string from proper names which is the best match
        key = process.extractOne(trial_name, proper_names)[0]

        # Use this to index into the dictionary and return the full dict for the best matching item
        single_clinic_dict = all_clinics_dict[key]

    except TypeError:

        # If a match cannot be made, potentially because proper_names is empty as all_clinics_dict is empty,
        # use descriptive placeholders

        single_clinic_dict = {
            'status': 'Not Found',        # Clinic Type
            'district': 'Not Found',      # District
            'funding': 'Not Found',       # Donor
            'pop': 'Not Found'            # Est.Pop.
        }

    return single_clinic_dict
